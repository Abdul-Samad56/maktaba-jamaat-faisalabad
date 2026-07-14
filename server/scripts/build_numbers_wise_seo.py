# -*- coding: utf-8 -*-
"""Extract name seeds from NUMBERS Wise G.xlsm and emit SEO data files."""
from __future__ import annotations

import json
import pathlib
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

XLSX = r"e:\IT\Facebook\NUMBERS Wise G.xlsm"
ROOT = pathlib.Path(r"e:\مکتبہ\پبلیشرز\Maktaba Jamaat e Islami Faisalabad")
SERVER_DATA = ROOT / "server" / "src" / "data"
CLIENT_SRC = ROOT / "client" / "src"
TMP = ROOT / "server" / "tmp_numbers"
TMP.mkdir(parents=True, exist_ok=True)

# Huge low-signal scrapes — only take Islam-filtered samples
HUGE_SHEETS = {"445", "159", "90+", "90plus"}
SKIP_SHEETS = {"Summary"}

# Prefer larger Facebook audience buckets first
SHEET_ORDER = [
    "9M", "8M", "7M", "6M", "5M", "4M", "3M", "2M", "1M",
    "900k", "800k", "700k", "600k", "500k", "400k", "300k", "200k", "100k",
    "90k", "80k", "70k", "60k", "50k", "40k", "30k", "20k", "10k",
    "90plus", "90+", "Sheet1", "445", "159",
]

CLIENT_CAP = 2500
HUGE_ISLAMIC_CAP = 300  # from huge sheets, Islamic only

ISLAMIC_HINTS = [
    "islam", "islami", "islamic", "quran", "quraan", "hadith", "hadees", "namaz",
    "salah", "mosque", "masjid", "maktaba", "jamaat", "ji ", "maududi", "maudoodi",
    "zakir", "naik", "naat", "deen", "sunnah", "sunnat", "madina", "makkah", "mecca",
    "madinah", "tafseer", "tafsir", "fiqh", "seerat", "seerah", "dua", "ayah",
    "prophet", "rasool", "rasul", "allah", "muslim", "ummah", "tabligh", "tableegh",
    "deoband", "barelvi", "ahle", "ahmad", "usmani", "israr", "tariq jameel",
    "tariq jamil", "minhaj", "khidmat", "tarjuman", "tafheem", "iqbal",
    "قرآن", "اسلام", "حدیث", "نماز", "مسجد", "مکتبہ", "جماعت", "مودودی", "ذاکر",
    "نائیک", "نعت", "دین", "سنت", "مدینہ", "مکہ", "تفسیر", "فقہ", "سیرت", "دعا",
    "رسول", "اللہ", "مسلم", "تبلیغ", "دیوبند", "بریلوی", "عثمانی", "اسرار",
    "طارق جمیل", "منہاج", "خدمت", "ترجمان", "تفہیم", "اقبال", "قصص رسول",
    "محمود الحسنات", "اقرأ", "حلیمہ", "زکریا", "bukhari", "muslim", "tirmizi",
    "ramadan", "رمضان", "hajj", "haj", "umrah", "عمرہ", "حج", "روحانی", "وظیفہ",
]

BUY_SELL_RE = re.compile(
    r"buy\s*(and|&|/)?\s*sell|for\s*sale|sale\s*and\s*purchase|sasta\s*bazar|"
    r"property|real\s*estate|cars?\s*(sale|market)|mobile\s*(sale|phone)|"
    r"memes?|coin\s*master|ikea|interior\s*design|true\s*love|soulmate|"
    r"trading\s*group|reposessed|repossessed",
    re.I,
)

NSFW_RE = re.compile(r"sexy|xx+|adult|porn|nude|hot\s*girl", re.I)


def clean(s: object) -> str:
    t = str(s or "").strip()
    t = re.sub(r"[\u200e\u200f\u202a-\u202e\ufeff]", "", t)
    t = re.sub(r"[\U0001F300-\U0001FAFF\u2600-\u27BF★☆♥❤◆●•]+", "", t)
    t = re.sub(r"\s+", " ", t).strip(" -_|./")
    return t


def is_junk(t: str) -> bool:
    if not t or len(t) < 2 or len(t) > 70:
        return True
    low = t.lower()
    if low.startswith("x1") or "href" in low:
        return True
    if "http://" in low or "https://" in low or "facebook.com" in low or "fb.com" in low:
        return True
    if low == "public" or low.startswith("public ·") or low.startswith("public."):
        return True
    if "members" in low or "posts a day" in low:
        return True
    if re.fullmatch(r"\d+[kKmM+]?", t):
        return True
    if NSFW_RE.search(t):
        return True
    if not re.search(r"[A-Za-z\u0600-\u06FF]", t):
        return True
    return False


def is_islamic(term: str) -> bool:
    low = term.casefold()
    return any(h in low or h in term for h in ISLAMIC_HINTS)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_vba=False)
    by_sheet: dict[str, int] = {}
    flat: list[dict] = []
    seen: set[str] = set()
    huge_islamic = 0

    for sn in SHEET_ORDER:
        if sn not in wb.sheetnames or sn in SKIP_SHEETS:
            continue
        ws = wb[sn]
        kept = 0
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            t = clean(row[1])
            if is_junk(t):
                continue
            # Prefer bookstore-useful names; drop obvious buy/sell noise unless Islamic
            islamic = is_islamic(t)
            if BUY_SELL_RE.search(t) and not islamic:
                continue

            if sn in HUGE_SHEETS:
                if not islamic:
                    continue
                if huge_islamic >= HUGE_ISLAMIC_CAP and len(flat) >= CLIENT_CAP:
                    break
                if not islamic:
                    continue

            key = t.casefold()
            if key in seen:
                continue
            seen.add(key)
            kept += 1
            if sn in HUGE_SHEETS and islamic:
                huge_islamic += 1
            flat.append({"term": t, "sheet": sn, "islamic": islamic})

            if len(flat) >= CLIENT_CAP and sn not in HUGE_SHEETS:
                # Still allow huge sheets to add Islamic-only until HUGE_ISLAMIC_CAP
                if sn not in HUGE_SHEETS:
                    break
        by_sheet[sn] = kept
        if len(flat) >= CLIENT_CAP + HUGE_ISLAMIC_CAP:
            break

    wb.close()

    # Cap final list
    if len(flat) > CLIENT_CAP:
        # Keep all Islamic first, then fill
        islamic_rows = [x for x in flat if x["islamic"]]
        other = [x for x in flat if not x["islamic"]]
        flat = (islamic_rows + other)[:CLIENT_CAP]

    featured = [x for x in flat if x["islamic"]][:150]
    if len(featured) < 40:
        featured = flat[:80]

    meta = {
        "source": XLSX,
        "sheets": by_sheet,
        "unique_terms": len(flat),
        "featured_count": len(featured),
        "sample": flat[:40],
    }
    (TMP / "_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (TMP / "_all_terms.json").write_text(json.dumps(flat, ensure_ascii=False), encoding="utf-8")

    SERVER_DATA.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedFrom": "NUMBERS Wise G.xlsm",
        "uniqueTerms": len(flat),
        "keywordsPerTerm": 200,
        "terms": flat,
        "featured": featured,
        "sitemapTerms": (featured + [x for x in flat if not x["islamic"]])[:1000],
    }
    (SERVER_DATA / "numbersWiseTerms.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )

    js_terms = json.dumps([t["term"] for t in flat], ensure_ascii=False)
    js_featured = json.dumps([t["term"] for t in featured], ensure_ascii=False)
    js = f"""/**
 * Auto-generated from NUMBERS Wise G.xlsm
 * Do not edit by hand — run server/scripts/build_numbers_wise_seo.py
 * Each term expands to 200 SEO keywords via expandRoadmapKeywords().
 */
export const NUMBERS_WISE_SOURCE = "NUMBERS Wise G.xlsm";
export const NUMBERS_WISE_KEYWORDS_PER_TERM = 200;
export const NUMBERS_WISE_TERM_COUNT = {len(flat)};

/** Curated unique group/person name seeds (member-count sheets). */
export const NUMBERS_WISE_TERMS = {js_terms};

/** Islam / bookstore-adjacent featured seeds. */
export const NUMBERS_WISE_FEATURED = {js_featured};
"""
    (CLIENT_SRC / "numbersWiseSeoData.js").write_text(js, encoding="utf-8")

    print(
        json.dumps(
            {
                "unique": len(flat),
                "featured": len(featured),
                "sheets": by_sheet,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
