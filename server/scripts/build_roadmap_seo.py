# -*- coding: utf-8 -*-
"""Extract seed terms from Road Map workbook and emit SEO data files."""
from __future__ import annotations

import json
import pathlib
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

XLSX = r"e:\IT\Facebook\Road Map and Famous Person List.xlsx"
ROOT = pathlib.Path(r"e:\مکتبہ\پبلیشرز\Maktaba Jamaat e Islami Faisalabad")
SERVER_DATA = ROOT / "server" / "src" / "data"
CLIENT_SRC = ROOT / "client" / "src"
TMP = ROOT / "server" / "tmp_roadmap"
TMP.mkdir(parents=True, exist_ok=True)

# Skip scrapes / instructions. Khyber "Area" sheet is mostly Facebook group titles.
SKIP_SHEETS = {"Road Map", "ڈیٹا", "Sheet1", "Area of cities khayber"}
USE_SHEETS = [
    "سرچ کے تجویز نام",
    "جو پرسن سرچ ہو گئے ہے۔",
    "Cities Name",
    "Area Name",
    "Name Hint",
]

JUNK = {
    "",
    "nan",
    "none",
    "sub type",
    "cities",
    "countries",
    "نام",
    "ںام",
    "نمبر شمار",
    "ممبرز تعداد",
    "computer",
    "latop",
    "laptop",
    "election pc",
    "election cell",
    "foundation",
    "s.no",
    "district name",
    "maktaba computer",
    "election cell computer",
    "public · 1 member",
}

# Ultra-generic / harmful / non-bookstore single tokens
GENERIC_EN = {
    "good", "new", "viral", "funny", "easy", "vote", "voice", "home", "house",
    "party", "memes", "latest", "trending", "team", "love", "like", "follow",
    "create", "speak", "about", "work", "need", "start", "official", "live",
    "post", "blog", "boss", "info", "local", "share", "zone", "part", "quiz",
    "offer", "picture", "friend", "ads", "off", "car", "dog", "goat", "cow",
    "club", "gem", "horse", "hub", "media", "social", "cup", "rose", "flowers",
    "animal", "tv", "single", "beautiful", "sweet", "bike", "services", "mall",
    "metro", "excel", "seo", "cyber", "sim", "css", "tiger", "fox", "rabbit",
    "snake", "frog", "eagle", "parrot", "duck", "art", "zero", "tips", "trick",
    "brand", "star", "fitness", "design", "outfit", "coat", "jacket", "luxury",
    "fashion", "nature", "water", "air", "earth", "sky", "linen", "fabric",
    "salon", "studio", "trend", "drama", "episode", "magic", "support", "award",
    "prize", "verify", "package", "internet", "student", "travel", "agent",
    "champion", "practice", "goal", "league", "game", "ludo", "prof", "sector",
    "dr", "line", "welcome", "night", "wish", "blessing", "care", "god",
    "union", "capital", "public", "child", "profit", "solar", "tex", "pizza",
    "galaxy", "generation", "novel", "tradition", "season", "all", "com", "net",
    "org", "gov", "portal", "pak", "kp", "ngo", "sell", "report", "ps", "pb",
    "jan", "ishq", "leak", "miss", "mr", "mrs", "madam", "mam", "alert",
    "imam", "community", "forum", "lovers", "today", "sunni", "opportunity",
    "sister", "brother", "savvy", "vibes", "true", "teen", "showbiz", "trust",
    "heroes", "legend", "point", "park", "aunty", "sexy", "xxzz", "bank",
    "store", "depo", "centre", "ltd", "clothe", "mill", "minister", "railway",
    "passport", "laboratory", "medical", "police", "rangers", "fia", "army",
    "news", "map", "mna", "invester", "invesment", "online", "earning", "shoe",
    "book", "electronic", "frage", "fan", "offical", "acters", "quotes", "jesus",
    "shadi", "rishta", "furniture", "soprt", "baryani", "butt", "malak",
    "nursery", "class 1", "ssc", "ba", "bs", "bachelor", "dae", "nursing",
    "lhv", "dispenser", "ot", "x-ray", "radiology", "ultrasound", "dental",
    "physiotherapy", "dcom", "dba", "accounting", "finance", "hrm", "textile",
    "plumber", "welding", "refrigeration", "cooking", "hotel", "driving", "nurse",
    "chinese", "mathematics", "hotal", "job", "criket", "women", "girls",
    "companies", "foundations", "mazahid", "shia", "hindu", "sikh", "petrolem",
    "hotals", "fast food", "languages", "collages", "transport", "industry",
    "property", "market", "farmer", "youth", "tourism", "trader", "factory",
    "properties", "university", "college", "hospitals", "tractor", "shareef",
    "islamic", "madaris", "school list", "all masalik", "peer/mureed",
}

NSFW_RE = re.compile(
    r"sexy|xxzz|مستی|گرم\s*لڑک|مردانہ کمزوری|جنسی",
    re.I,
)

JUNK_RE = re.compile(
    r"^(sub\s*type|cities|countries|نام|ںام|نمبر|ممبرز|latop|laptop|computer|s\.?no)$",
    re.I,
)

FB_GROUP_RE = re.compile(
    r"buy\s*(and|&)\s*sell|free\s*advertis|members|friends who like|"
    r"fan\s*club|official\s*group|property\s*(agents|business|community)|"
    r"houses for sale|rental property|investment property",
    re.I,
)


def clean(s: object) -> str:
    t = str(s or "").strip()
    t = re.sub(r"[\u200e\u200f\u202a-\u202e\ufeff]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    # strip emoji / fancy unicode symbols that break SEO
    t = re.sub(r"[\U0001F300-\U0001FAFF\u2600-\u27BF★☆♥❤◆●•]+", "", t)
    t = re.sub(r"\s+", " ", t).strip(" -_|./")
    return t


def is_junk(t: str) -> bool:
    if not t or len(t) < 2:
        return True
    if t.lower() in JUNK or t.lower() in GENERIC_EN:
        return True
    if JUNK_RE.match(t):
        return True
    if re.fullmatch(r"\d+", t):
        return True
    if len(t) > 55:
        return True
    low = t.lower()
    if "http://" in low or "https://" in low or "facebook.com" in low or "fb.com" in low:
        return True
    if low.startswith("x1") or "href" in low:
        return True
    if NSFW_RE.search(t) or FB_GROUP_RE.search(t):
        return True
    if "سرچ" in t and len(t) > 25:
        return True
    if "گروپ کو جائن" in t or "لسٹ چاہیے" in t:
        return True
    if low in {"list", "list of domain", "done", "the name of brands list"}:
        return True
    # pure punctuation / emoji leftovers
    if not re.search(r"[A-Za-z\u0600-\u06FF]", t):
        return True
    return False


def sheet_kind(sn: str) -> str:
    if "Cities" in sn or sn == "Cities Name":
        return "city"
    if "Area" in sn:
        return "area"
    if sn == "Name Hint":
        return "category"
    if "تجویز" in sn:
        return "suggested"
    if "پرسن" in sn:
        return "person"
    return "term"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    by_sheet: dict[str, list[str]] = {}
    flat: list[dict] = []
    seen: set[str] = set()

    for sn in USE_SHEETS:
        if sn not in wb.sheetnames or sn in SKIP_SHEETS:
            continue
        ws = wb[sn]
        terms: list[str] = []
        local: set[str] = set()
        kind = sheet_kind(sn)
        person_mode = "پرسن" in sn
        name_hint = sn == "Name Hint"
        cities_mode = sn == "Cities Name"

        for row in ws.iter_rows(values_only=True):
            cells = list(row) if row else []
            if person_mode and cells:
                candidates = [cells[0]]
            elif name_hint:
                candidates = [c for c in cells if c and len(clean(c)) <= 40]
            elif cities_mode:
                # EN + UR name columns typically first two meaningful cells
                candidates = cells[:4]
            else:
                candidates = cells

            for cell in candidates:
                t = clean(cell)
                if is_junk(t):
                    continue
                key = t.casefold()
                if key in local:
                    continue
                local.add(key)
                terms.append(t)
                if key not in seen:
                    seen.add(key)
                    flat.append({"term": t, "sheet": sn, "kind": kind})

        by_sheet[sn] = terms
        (TMP / f"{sn[:50].replace('/', '-')}.json").write_text(
            json.dumps(terms, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    wb.close()

    FEATURED_HINTS = [
        "jamaat", "islami", "islamic", "maududi", "maudoodi", "naeem", "siraj",
        "qazi", "munawar", "tufail", "mushtaq", "tariq jameel", "tariq jamil",
        "israr", "usmani", "taqi", "al khidmat", "khidmat", "tarjuman", "tafheem",
        "quran", "حدیث", "قرآن", "جماعت", "اسلامی", "مودودی", "نعیم", "سراج",
        "قاضی", "منور", "طفیل", "مشتاق", "tableegi", "تبلیغی", "maktaba", "مکتبہ",
        "madina", "faisalabad", "فیصل", "lahore", "لاہور", "karachi", "کراچی",
        "islamabad", "اسلام آباد", "rawalpindi", "multan", "peshawar", "gujranwala",
        "sialkot", "bano", "بنو", "iqbal", "اقبال", "bukhari", "بخاری", "namaz",
        "نماز", "zakir", "ذاکر", "farhat", "فرحت", "minhaj", "منہاج", "دعوت",
        "دیوبند", "بریلوی", "seerat", "سیرت", "hadith", "fiqh", "فقہ",
    ]

    def is_featured(term: str) -> bool:
        low = term.casefold()
        return any(h in low or h in term for h in FEATURED_HINTS)

    featured = [x for x in flat if is_featured(x["term"])]
    cities = [x for x in flat if x["kind"] == "city"]
    persons = [x for x in flat if x["kind"] in ("person", "suggested")]
    areas = [x for x in flat if x["kind"] == "area"]

    meta = {
        "source": XLSX,
        "sheets": {k: len(v) for k, v in by_sheet.items()},
        "unique_terms": len(flat),
        "persons_suggested": len(persons),
        "cities": len(cities),
        "areas": len(areas),
        "featured_count": len(featured),
        "sample": flat[:40],
    }
    (TMP / "_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (TMP / "_all_terms.json").write_text(json.dumps(flat, ensure_ascii=False), encoding="utf-8")

    SERVER_DATA.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedFrom": "Road Map and Famous Person List.xlsx",
        "uniqueTerms": len(flat),
        "keywordsPerTerm": 200,
        "terms": flat,
        "featured": featured[:200],
        "cities": cities,
        "sitemapPersons": persons[:800],
        "sitemapCities": cities,
        "sitemapAreas": areas[:300],
    }
    (SERVER_DATA / "roadmapTerms.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )

    js_terms = json.dumps([t["term"] for t in flat], ensure_ascii=False)
    js_featured = json.dumps([t["term"] for t in featured[:120]], ensure_ascii=False)
    js_cities = json.dumps([t["term"] for t in cities], ensure_ascii=False)
    js = f"""/**
 * Auto-generated from Road Map and Famous Person List.xlsx
 * Do not edit by hand — run server/scripts/build_roadmap_seo.py
 * Each term expands to 200 SEO keywords via expandRoadmapKeywords().
 */
export const ROADMAP_SOURCE = "Road Map and Famous Person List.xlsx";
export const ROADMAP_KEYWORDS_PER_TERM = 200;
export const ROADMAP_TERM_COUNT = {len(flat)};

/** All unique seed terms (persons, cities, areas, categories). */
export const ROADMAP_TERMS = {js_terms};

/** Bookstore / Islam-adjacent featured seeds for homepage & meta. */
export const ROADMAP_FEATURED = {js_featured};

/** City seeds from Cities Name sheet. */
export const ROADMAP_CITIES = {js_cities};
"""
    (CLIENT_SRC / "roadmapSeoData.js").write_text(js, encoding="utf-8")

    print(
        json.dumps(
            {
                "unique": len(flat),
                "featured": len(featured),
                "cities": len(cities),
                "persons": len(persons),
                "sheets": meta["sheets"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
