# -*- coding: utf-8 -*-
"""Extract name seeds from NAME ID.xlsm and emit SEO data files."""
from __future__ import annotations

import json
import pathlib
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

XLSX = r"e:\IT\Facebook\NAME ID.xlsm"
ROOT = pathlib.Path(r"e:\مکتبہ\پبلیشرز\Maktaba Jamaat e Islami Faisalabad")
SERVER_DATA = ROOT / "server" / "src" / "data"
CLIENT_SRC = ROOT / "client" / "src"
TMP = ROOT / "server" / "tmp_nameid"
TMP.mkdir(parents=True, exist_ok=True)

# Setting/Filter largely duplicate لسٹ نمبر1 — skip duplicates via seen set
SHEET_ORDER = [
    "سرچ نام",           # curated EN+UR search names (priority)
    "اردو",               # Urdu profile names
    "تجویز کے لے نام",   # suggested first names
    "لسٹ نمبر1",
    "Adding From This Point",
    # Setting/Filter ≈ duplicates of لسٹ نمبر1
]

CLIENT_CAP = 2500
# Finish these fully before enforcing CLIENT_CAP
PRIORITY_SHEETS = {"سرچ نام", "اردو"}

ISLAMIC_HINTS = [
    "islam", "islami", "islamic", "quran", "hadith", "namaz", "mosque", "masjid",
    "maktaba", "jamaat", "madina", "makkah", "allah", "muslim", "fatima", "amina",
    "ayesha", "aisha", "maryam", "khadija", "zainab", "hafsa", "safiya", "ruqayya",
    "sumayya", "asma", "iqra", "noor", "iman", "eeman", "muhammad", "ahmad",
    "ali", "hassan", "hussain", "umar", "usman", "bilal", "hamza", "yusuf",
    "اسلام", "قرآن", "فاطمہ", "آمنہ", "عائشہ", "مریم", "خدیجہ", "زینب", "حفصہ",
    "اقرا", "نور", "ایمان", "محمد", "احمد", "علی", "حسن", "حسین", "جنت", "مدینہ",
    "مکتبہ", "جماعت", "مدرسہ", "نعت", "دین",
]

HEADER_RE = re.compile(
    r"^(no|name|add friend|follow|join|visit|setting|filter)$",
    re.I,
)
NSFW_RE = re.compile(r"sexy|xx+|adult|porn|nude|hot\s*girl|doll\b", re.I)
PHONE_RE = re.compile(r"^[\d\s+\-()]{7,}$")


def clean(s: object) -> str:
    t = str(s or "").strip()
    t = re.sub(r"[\u200e\u200f\u202a-\u202e\ufeff]", "", t)
    # strip wrapping quotes / RTL marks leftovers
    t = t.strip(" \"'`")
    t = re.sub(r"[\U0001F300-\U0001FAFF\u2600-\u27BF★☆♥❤◆●•]+", "", t)
    # remove parenthetical nicknames noise for cleaner seeds when whole cell is messy
    t = re.sub(r"\s+", " ", t).strip(" -_|./")
    return t


def clean_person_name(t: str) -> str:
    """Prefer primary name before nickname in parentheses."""
    t = clean(t)
    if not t:
        return ""
    # "Iqra Doll (iqra doll)" -> keep left part if reasonable
    m = re.match(r"^([^((]+)", t)
    if m:
        left = m.group(1).strip()
        if len(left) >= 2:
            t = left
    # strip leading special chars
    t = re.sub(r"^[^A-Za-z\u0600-\u06FF]+", "", t).strip()
    return t


def is_junk(t: str) -> bool:
    if not t or len(t) < 2 or len(t) > 60:
        return True
    if HEADER_RE.match(t):
        return True
    if PHONE_RE.match(t):
        return True
    low = t.lower()
    if low.startswith("x1") or "href" in low:
        return True
    if "http://" in low or "https://" in low or "facebook.com" in low or "fb.com" in low:
        return True
    if low.startswith("@") or "add friend" in low or "lives in" in low:
        return True
    if NSFW_RE.search(t):
        return True
    if not re.search(r"[A-Za-z\u0600-\u06FF]", t):
        return True
    # too generic single letters / numbers only
    if re.fullmatch(r"\d+", t):
        return True
    return False


def is_islamic(term: str) -> bool:
    low = term.casefold()
    return any(h in low or h in term for h in ISLAMIC_HINTS)


def add_term(flat, seen, term: str, sheet: str, kind: str = "name"):
    t = clean_person_name(term) if kind == "person" else clean(term)
    if kind == "person":
        t = clean_person_name(term)
    else:
        t = clean(term)
    if is_junk(t):
        return False
    key = t.casefold()
    if key in seen:
        return False
    seen.add(key)
    flat.append(
        {
            "term": t,
            "sheet": sheet,
            "kind": kind,
            "islamic": is_islamic(t),
        }
    )
    return True


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_vba=False)
    by_sheet: dict[str, int] = {}
    flat: list[dict] = []
    seen: set[str] = set()

    for sn in SHEET_ORDER:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        kept = 0

        for row in ws.iter_rows(values_only=True):
            cells = list(row) if row else []
            if not cells:
                continue

            if sn == "سرچ نام":
                # EN name + UR name
                for idx in (0, 2):
                    if idx < len(cells) and add_term(flat, seen, cells[idx], sn, "search"):
                        kept += 1
            elif sn == "تجویز کے لے نام":
                if add_term(flat, seen, cells[0], sn, "suggested"):
                    kept += 1
            elif sn == "اردو":
                if len(cells) > 1 and add_term(flat, seen, cells[1], sn, "person"):
                    kept += 1
            else:
                # FB profile lists — name in column B
                if len(cells) > 1 and add_term(flat, seen, cells[1], sn, "person"):
                    kept += 1

            if len(flat) >= CLIENT_CAP and sn not in PRIORITY_SHEETS:
                break

        by_sheet[sn] = kept
        if len(flat) >= CLIENT_CAP and sn not in PRIORITY_SHEETS:
            break

    wb.close()

    # Prefer search-sheet + Islamic names first in final cap
    if len(flat) > CLIENT_CAP:
        search_rows = [x for x in flat if x["kind"] == "search"]
        islamic_rows = [x for x in flat if x["islamic"] and x["kind"] != "search"]
        other = [x for x in flat if not x["islamic"] and x["kind"] != "search"]
        flat = (search_rows + islamic_rows + other)[:CLIENT_CAP]

    featured = [x for x in flat if x["kind"] == "search"][:80]
    featured += [x for x in flat if x["islamic"] and x not in featured][:80]
    if len(featured) < 40:
        featured = flat[:100]
    featured = featured[:160]

    meta = {
        "source": XLSX,
        "sheets": by_sheet,
        "unique_terms": len(flat),
        "featured_count": len(featured),
        "sample": flat[:40],
    }
    (TMP / "_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (TMP / "_all_terms.json").write_text(json.dumps(flat, ensure_ascii=False), encoding="utf-8")

    SERVER_DATA.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedFrom": "NAME ID.xlsm",
        "uniqueTerms": len(flat),
        "keywordsPerTerm": 200,
        "terms": flat,
        "featured": featured,
        "sitemapTerms": (featured + [x for x in flat if x not in featured])[:1000],
    }
    (SERVER_DATA / "nameIdTerms.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )

    js_terms = json.dumps([t["term"] for t in flat], ensure_ascii=False)
    js_featured = json.dumps([t["term"] for t in featured], ensure_ascii=False)
    js = f"""/**
 * Auto-generated from NAME ID.xlsm
 * Do not edit by hand — run server/scripts/build_name_id_seo.py
 * Each term expands to 200 SEO keywords via expandRoadmapKeywords().
 */
export const NAME_ID_SOURCE = "NAME ID.xlsm";
export const NAME_ID_KEYWORDS_PER_TERM = 200;
export const NAME_ID_TERM_COUNT = {len(flat)};

/** Curated unique name seeds from NAME ID workbook. */
export const NAME_ID_TERMS = {js_terms};

/** Priority search / Islamic featured names. */
export const NAME_ID_FEATURED = {js_featured};
"""
    (CLIENT_SRC / "nameIdSeoData.js").write_text(js, encoding="utf-8")

    print(
        json.dumps(
            {
                "unique": len(flat),
                "featured": len(featured),
                "sheets": by_sheet,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
