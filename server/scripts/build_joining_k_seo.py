# -*- coding: utf-8 -*-
"""Extract name seeds from Joining K data.xlsm and emit SEO data files."""
from __future__ import annotations

import json
import pathlib
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

XLSX = r"e:\IT\Facebook\Joining K data.xlsm"
ROOT = pathlib.Path(r"e:\مکتبہ\پبلیشرز\Maktaba Jamaat e Islami Faisalabad")
SERVER_DATA = ROOT / "server" / "src" / "data"
CLIENT_SRC = ROOT / "client" / "src"
TMP = ROOT / "server" / "tmp_joining"
TMP.mkdir(parents=True, exist_ok=True)

# Sheet "k" has name in column C (index 2); others usually column B (index 1)
NAME_COL = {
    "k": 2,
}

SHEET_ORDER = [
    "k",
    "1K",
    "2K",
    "3K",
    "4K",
    "5K",
    "6K",
    "7 5h",
    "8 2h",
    "9 low 200",
    "Sheet1",
    "PROPERTY",  # large property scrape — filtered heavily
]

CLIENT_CAP = 2500
PROPERTY_CAP = 400

ISLAMIC_HINTS = [
    "islam", "islami", "islamic", "quran", "quraan", "hadith", "hadees", "namaz",
    "salah", "mosque", "masjid", "maktaba", "jamaat", "maududi", "maudoodi",
    "zakir", "naik", "naat", "deen", "sunnah", "sunnat", "madina", "makkah",
    "madinah", "tafseer", "tafsir", "fiqh", "seerat", "dua", "prophet", "rasool",
    "allah", "muslim", "tabligh", "tableegh", "deoband", "barelvi", "usmani",
    "israr", "tariq jameel", "minhaj", "khidmat", "tarjuman", "tafheem", "iqbal",
    "جامعہ", "طلباء", "اسلامیہ", "قرآن", "اسلام", "حدیث", "نماز", "مسجد", "مکتبہ",
    "جماعت", "مودودی", "ذاکر", "نعت", "دین", "سنت", "مدینہ", "مکہ", "تفسیر", "فقہ",
    "سیرت", "دعا", "رسول", "اللہ", "مسلم", "تبلیغ", "دیوبند", "بریلوی", "عثمانی",
    "اسرار", "طارق جمیل", "منہاج", "خدمت", "ترجمان", "تفہیم", "اقبال", "حفظ",
    "رمضان", "حج", "عمرہ", "روحانی", "امدادیہ", "educators", "school", "college",
    "university", "madrasa", "madrasah", "مدرسہ", "کتب", "کتاب", "book",
]

LOCAL_HINTS = [
    "faisalabad", "fsd", "فیصل آباد", "chiniot", "چنیوٹ", "madina town",
    "مدینہ ٹاؤن", "people", "gulberg", "گلبرگ", "jaranwala", "جڑانوالہ",
    "sargodha road", "canal", "mansoorabad", "منصورآباد", "lyallpur", "لائل پور",
]

BUY_SELL_RE = re.compile(
    r"buy\s*(and|&|/)?\s*sell|for\s*sale|sale\s*and\s*purchase|sasta\s*bazar|"
    r"property|real\s*estate|cars?\s*(sale|market)|mobile\s*(sale|phone)|"
    r"memes?|biker\s*boys|motor\s*market|foods?\s*lover|foodies|"
    r"industrial\s*property|wapda\s*city\s*property|"
    r"خرید\s*و\s*فروخت|خرید\s*فروخت|بیچو|فروخت",
    re.I,
)

NSFW_RE = re.compile(r"sexy|xx+|adult|porn|nude|hot\s*girl", re.I)
HEADER_RE = re.compile(r"^(no|fb group name|number of members|join|visit)$", re.I)


def clean(s: object) -> str:
    t = str(s or "").strip()
    t = re.sub(r"[\u200e\u200f\u202a-\u202e\ufeff]", "", t)
    t = re.sub(r"[\U0001F300-\U0001FAFF\u2600-\u27BF★☆♥❤◆●•]+", "", t)
    t = re.sub(r"\s+", " ", t).strip(" -_|./")
    return t


def is_junk(t: str) -> bool:
    if not t or len(t) < 2 or len(t) > 80:
        return True
    if HEADER_RE.match(t):
        return True
    low = t.lower()
    if low.startswith("x1") or "href" in low:
        return True
    if "http://" in low or "https://" in low or "facebook.com" in low or "fb.com" in low:
        return True
    if low == "public" or low.startswith("public ·") or low.startswith("public."):
        return True
    if "members" in low or "posts a" in low or "friends are members" in low:
        return True
    if re.fullmatch(r"\d+[kKmM+]?", t) or re.fullmatch(r"\d+", t):
        return True
    if NSFW_RE.search(t):
        return True
    if not re.search(r"[A-Za-z\u0600-\u06FF]", t):
        return True
    return False


def is_islamic(term: str) -> bool:
    low = term.casefold()
    return any(h in low or h in term for h in ISLAMIC_HINTS)


def is_local(term: str) -> bool:
    low = term.casefold()
    return any(h in low or h in term for h in LOCAL_HINTS)


def name_from_row(sn: str, row) -> str:
    idx = NAME_COL.get(sn, 1)
    if not row or len(row) <= idx:
        # fallback: first non-url text cell
        for cell in row or []:
            t = clean(cell)
            if t and not t.lower().startswith("http") and not is_junk(t):
                return t
        return ""
    return clean(row[idx])


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_vba=False)
    by_sheet: dict[str, int] = {}
    flat: list[dict] = []
    seen: set[str] = set()
    property_kept = 0

    for sn in SHEET_ORDER:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        kept = 0
        for row in ws.iter_rows(values_only=True):
            t = name_from_row(sn, row)
            if is_junk(t):
                continue

            islamic = is_islamic(t)
            local = is_local(t)

            if sn == "PROPERTY":
                # Islamic / education / clean locality only
                if islamic:
                    pass
                elif local and not BUY_SELL_RE.search(t):
                    if property_kept >= PROPERTY_CAP:
                        continue
                else:
                    continue
                property_kept += 1
            else:
                # Skip buy/sell noise unless Islamic
                if BUY_SELL_RE.search(t) and not islamic:
                    continue

            key = t.casefold()
            if key in seen:
                continue
            seen.add(key)
            kept += 1
            flat.append(
                {
                    "term": t,
                    "sheet": sn,
                    "islamic": islamic,
                    "local": local,
                }
            )

            if len(flat) >= CLIENT_CAP:
                break
        by_sheet[sn] = kept
        if len(flat) >= CLIENT_CAP:
            break

    wb.close()

    if len(flat) > CLIENT_CAP:
        islamic_rows = [x for x in flat if x["islamic"]]
        local_rows = [x for x in flat if x["local"] and not x["islamic"]]
        other = [x for x in flat if not x["islamic"] and not x["local"]]
        flat = (islamic_rows + local_rows + other)[:CLIENT_CAP]

    featured = [x for x in flat if x["islamic"]][:120]
    featured += [
        x
        for x in flat
        if x["local"] and not x["islamic"] and not BUY_SELL_RE.search(x["term"])
    ][:40]
    if len(featured) < 40:
        featured = flat[:80]

    meta = {
        "source": XLSX,
        "sheets": by_sheet,
        "unique_terms": len(flat),
        "featured_count": len(featured),
        "sample": flat[:40],
    }
    (TMP / "_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (TMP / "_all_terms.json").write_text(json.dumps(flat, ensure_ascii=False), encoding="utf-8")

    SERVER_DATA.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedFrom": "Joining K data.xlsm",
        "uniqueTerms": len(flat),
        "keywordsPerTerm": 200,
        "terms": flat,
        "featured": featured,
        "sitemapTerms": (featured + [x for x in flat if x not in featured])[:1000],
    }
    (SERVER_DATA / "joiningKTerms.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )

    js_terms = json.dumps([t["term"] for t in flat], ensure_ascii=False)
    js_featured = json.dumps([t["term"] for t in featured], ensure_ascii=False)
    js = f"""/**
 * Auto-generated from Joining K data.xlsm
 * Do not edit by hand — run server/scripts/build_joining_k_seo.py
 * Each term expands to 200 SEO keywords via expandRoadmapKeywords().
 */
export const JOINING_K_SOURCE = "Joining K data.xlsm";
export const JOINING_K_KEYWORDS_PER_TERM = 200;
export const JOINING_K_TERM_COUNT = {len(flat)};

/** Curated unique group/name seeds from Joining K workbook. */
export const JOINING_K_TERMS = {js_terms};

/** Islam / Faisalabad-local featured seeds. */
export const JOINING_K_FEATURED = {js_featured};
"""
    (CLIENT_SRC / "joiningKSeoData.js").write_text(js, encoding="utf-8")

    print(
        json.dumps(
            {
                "unique": len(flat),
                "featured": len(featured),
                "sheets": by_sheet,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
